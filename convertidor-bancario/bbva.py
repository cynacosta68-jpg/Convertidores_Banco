"""
Conversor: Otros Bancos (input) → Formato BBVA (output)

- Extrae del archivo de input los registros cuyo CBU pertenece al BBVA (017)
  usando el parser robusto de `otros_bancos` (soporta líneas corridas)
- SEPARA las personas jurídicas, que el BBVA rechaza: salen en un archivo
  aparte y el archivo principal se renumera y recalcula sin ellas
- Cruza con el Maestro Excel para obtener nombres formateados
- Genera archivo TXT de 250 caracteres/línea según modelo BBVA

Estructura del archivo de SALIDA (cada línea = 250 chars + CRLF):
  ┌─ 211 (Cabecera)       × 1
  ├─ 221 (Detalle dato)   ┐
  ├─ 222 (Detalle nombre) │ × N beneficiarios
  ├─ 223 (Detalle prov.)  │
  ├─ 224 (Detalle ref.)   ┘
  └─ 291 (Pie / totales)  × 1
"""

import io
from datetime import datetime
from openpyxl import load_workbook

try:
    from convertidores import otros_bancos
except ImportError:                      # ejecución con módulos planos
    import otros_bancos


# ══════════════════════════════════════════════════════════════
# CONFIGURACIÓN POR DEFECTO
# ══════════════════════════════════════════════════════════════
DEFAULTS = {
    "servicio":      "001261",
    "cod_banco":     "0017",
    "cta_debito":    "0097740103302327",
    "concepto":      "HONORARIOS",
    "moneda":        "ARS",
    "flag":          "0",
    "banco_nombre":  "FRANCES",
    "empresa":       "COLEGIO MEDICO DEL SUR DEL CHUBUT",
    "cod_extra":     "20",
    "provincia":     "CHUBUT",
    "referencia":    "LIQUIDACION",
    "tipo_cuenta":   "1",       # 1=CA, 0=CC
    # ── Separación de personas jurídicas ──
    "excluir_juridicas": True,
    "columna_juridica":  None,    # nombre de columna del maestro (opcional)
    "valores_juridica":  ("J", "JURIDICA", "JURÍDICA", "PJ", "EMPRESA", "SI", "S"),
}

LINE_LEN = 250

# Prefijos de CUIT/CUIL según AFIP.
#   20 / 23 / 24 / 25 / 26 / 27 → persona física
#   30 / 33 / 34                → persona jurídica
PREFIJOS_JURIDICA = ("30", "33", "34")
PREFIJOS_FISICA   = ("20", "23", "24", "25", "26", "27")


# ══════════════════════════════════════════════════════════════
# FUNCIONES DE APOYO
# ══════════════════════════════════════════════════════════════
def pad(texto: str, n: int) -> str:
    return texto[:n].ljust(n)


def cargar_maestro(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    maestro = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cbu_raw = str(row[idx["CBU"]] or "").strip()
        if not cbu_raw:
            continue
        maestro[cbu_raw] = {
            "nombre":     str(row[idx["NombreApellido"]] or "").strip(),
            "nro_cuenta": str(row[idx["NroCuenta"]] or "").strip(),
            "tipo_cobro": str(row[idx["TipoCobro"]] or "").strip(),
            # Fila completa, para poder consultar cualquier columna extra
            # (por ejemplo una que marque persona física / jurídica).
            "_fila": {h: row[i] for i, h in enumerate(headers) if h},
        }
    wb.close()
    return maestro


def parsear_fila_input(linea: str) -> dict:
    """
    Adaptador: delega en el parser robusto de `otros_bancos`.

    Antes esta función cortaba la línea por posiciones fijas, lo que
    devolvía importes truncados y CUITs corridos cuando el registro no
    medía exactamente 209 caracteres. Ver otros_bancos.py.
    """
    reg = otros_bancos.parsear_linea(linea)
    return {
        "cbu":      reg["cbu"],
        "importe":  reg["importe"],
        "nombre":   reg["nombre"],
        "tipo_doc": reg["tipo_doc"],
        "cuil":     reg["cuit"],
        "avisos":   reg["avisos"],
    }


def formatear_nombre_desde_maestro(nombre_maestro: str) -> str:
    partes = nombre_maestro.split()
    if not partes:
        return nombre_maestro
    apellidos = []
    nombres = []
    for p in partes:
        if p.isupper() and len(p) > 1:
            apellidos.append(p)
        else:
            nombres.append(p)
    if apellidos and nombres:
        nombre_fmt = " ".join(n.capitalize() if n.islower() else n for n in nombres)
        return f"{' '.join(apellidos)}, {nombre_fmt}"
    return nombre_maestro


def formatear_nombre_desde_input(nombre_input: str) -> str:
    if "," in nombre_input:
        ape, nom = nombre_input.split(",", 1)
        nom_title = " ".join(w.capitalize() for w in nom.strip().split())
        return f"{ape.strip()}, {nom_title}"
    return nombre_input


def es_persona_juridica(cuit: str, info: dict = None, cfg: dict = None) -> tuple:
    """
    Determina si el beneficiario es una persona jurídica.

    Criterio 1 (opcional): una columna del maestro marcada en cfg["columna_juridica"].
    Criterio 2 (por defecto): el prefijo del CUIT, que es regla de AFIP.

    Devuelve (es_juridica: bool, motivo: str).
    """
    cfg = cfg or DEFAULTS

    col = cfg.get("columna_juridica")
    if col and info and info.get("_fila"):
        valor = info["_fila"].get(col)
        if valor is not None and str(valor).strip():
            v = str(valor).strip().upper()
            valores = tuple(x.upper() for x in cfg.get("valores_juridica", ()))
            if v in valores:
                return True, f"maestro: columna '{col}' = '{valor}'"
            return False, f"maestro: columna '{col}' = '{valor}'"

    pref = (cuit or "").strip()[:2]
    if pref in PREFIJOS_JURIDICA:
        return True, f"CUIT {cuit} (prefijo {pref} = jurídica)"
    if pref in PREFIJOS_FISICA:
        return False, f"CUIT {cuit} (prefijo {pref} = física)"
    return False, f"CUIT {cuit} (prefijo {pref} desconocido: se trata como física)"


def buscar_en_maestro(maestro: dict, cbu: str):
    cbu_limpio = cbu.strip()
    if cbu_limpio in maestro:
        return maestro[cbu_limpio]
    for k, v in maestro.items():
        if k.strip() == cbu_limpio:
            return v
    return None


# ══════════════════════════════════════════════════════════════
# GENERADORES DE REGISTROS
# ══════════════════════════════════════════════════════════════
def build_header(fecha: str, cfg: dict) -> str:
    h = (
        "211"
        + cfg["servicio"]
        + fecha + fecha
        + cfg["cod_banco"]
        + pad(cfg["cta_debito"], 16)
        + pad(cfg["concepto"], 10)
        + cfg["moneda"]
        + cfg["flag"]
        + pad(cfg["banco_nombre"], 12)
        + pad(cfg["empresa"], 36)
        + pad(cfg["cod_extra"], 3)
    )
    return pad(h, LINE_LEN)


def build_221(seq: int, item: dict, fecha: str, cfg: dict) -> str:
    return pad(
        "221" + cfg["servicio"] + "  "
        + str(seq).zfill(18) + "    "
        + cfg["tipo_cuenta"]
        + item["cbu"]
        + str(item["importe"]).zfill(25)
        + "      " + fecha
        + item["cuil"].zfill(15),
        LINE_LEN
    )


def build_222(seq: int, nombre: str, cfg: dict) -> str:
    return pad(
        "222" + cfg["servicio"] + "  "
        + str(seq).zfill(18) + "    "
        + nombre,
        LINE_LEN
    )


def build_223(seq: int, cfg: dict) -> str:
    return pad(
        "223" + cfg["servicio"] + "  "
        + str(seq).zfill(18) + "    "
        + " " * 36 + cfg["provincia"],
        LINE_LEN
    )


def build_224(seq: int, cfg: dict) -> str:
    return pad(
        "224" + cfg["servicio"] + "  "
        + str(seq).zfill(18) + "    "
        + cfg["referencia"],
        LINE_LEN
    )


def build_footer(total_importe: int, cant_benef: int, total_lineas: int, cfg: dict) -> str:
    return pad(
        "291" + cfg["servicio"]
        + str(total_importe).zfill(15)
        + str(cant_benef).zfill(8) + "00"
        + str(total_lineas).zfill(8),
        LINE_LEN
    )


# ══════════════════════════════════════════════════════════════
# EXPORTACIÓN DE LAS PERSONAS JURÍDICAS SEPARADAS
# ══════════════════════════════════════════════════════════════
def _lineas_origen(registros: list) -> str:
    """
    Reconstruye las líneas en FORMATO DE ORIGEN (modelo estándar 209).

    Sirve para reprocesarlas por otro circuito sin tener que volver a
    tocar el archivo original a mano.
    """
    if not registros:
        return ""
    return "\r\n".join(otros_bancos.construir_linea(r) for r in registros) + "\r\n"


def _texto_excel(valor: str) -> str:
    """
    Fuerza a Excel a tratar el valor como TEXTO.

    Sin esto, un CBU como 0170097820000033043850 se abre como número, pierde
    el cero inicial y se muestra en notación científica (1,7E+20), quedando
    inservible para copiar al homebanking.
    """
    return f'="{valor}"'


def _csv_juridicas(registros: list) -> bytes:
    """Detalle legible en Excel (separador ';', UTF-8 con BOM)."""
    filas = ["CUIT;RazonSocial;NombreMaestro;CBU;Importe;Motivo"]
    for r in registros:
        importe = f"{r['importe'] // 100},{r['importe'] % 100:02d}"
        filas.append(";".join([
            _texto_excel(r["cuit"]),
            r["nombre"].replace(";", " "),
            str(r.get("nombre_maestro", "")).replace(";", " "),
            _texto_excel(r["cbu"]),
            importe,
            r.get("motivo_exclusion", "").replace(";", " "),
        ]))
    return ("\r\n".join(filas) + "\r\n").encode("utf-8-sig")


# ══════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════
def convertir(input_bytes: bytes, maestro_bytes: bytes,
              fecha_proceso: str = None, config: dict = None) -> dict:
    """
    Ejecuta la conversión completa.

    Parámetros:
        input_bytes:   bytes del archivo Otros_Bancos (input).txt
        maestro_bytes: bytes del archivo Maestro_datos_BBVA.xlsx
        fecha_proceso: str AAAAMMDD (default: hoy)
        config:        dict con valores de configuración (opcional)

    Retorna dict con:
        "contenido":      str con el archivo generado
        "beneficiarios":  int
        "total_lineas":   int
        "importe_total":  float
        "fecha":          str
        "no_maestro":     list de tuplas (cbu, nombre)
        "nombre_archivo": str sugerido
    """
    cfg = {**DEFAULTS, **(config or {})}
    if not fecha_proceso:
        fecha_proceso = datetime.today().strftime("%Y%m%d")

    maestro = cargar_maestro(maestro_bytes)

    # ── Lectura robusta del archivo Otros Bancos ──────────────────
    # Se filtra por código de banco del CBU (017 = BBVA). El parser
    # detecta la codificación real y reacomoda las líneas corridas.
    regs, analisis = otros_bancos.registros_para(
        input_bytes, incluir=[otros_bancos.COD_BBVA]
    )

    # ── Separar personas jurídicas (el BBVA las rechaza) ──────────
    juridicas = []
    aptos = []
    for r in regs:
        info = buscar_en_maestro(maestro, r["cbu"])
        if cfg["excluir_juridicas"]:
            es_jur, motivo = es_persona_juridica(r["cuit"], info, cfg)
            if es_jur:
                r = dict(r)
                r["motivo_exclusion"] = motivo
                r["nombre_maestro"] = (info or {}).get("nombre", "")
                juridicas.append(r)
                continue
        aptos.append(r)

    filas = [{
        "cbu":      r["cbu"],
        "importe":  r["importe"],
        "nombre":   r["nombre"],
        "tipo_doc": r["tipo_doc"],
        "cuil":     r["cuit"],
        "avisos":   r["avisos"],
    } for r in aptos]

    if not filas:
        return {
            "contenido": "",
            "beneficiarios": 0,
            "error": (
                f"No quedaron registros para el archivo BBVA. "
                f"Se leyeron {len(regs)} registro(s) del banco {otros_bancos.COD_BBVA} "
                f"y {len(juridicas)} fueron separados por ser personas jurídicas."
                if regs else
                f"No se encontraron registros del banco {otros_bancos.COD_BBVA} "
                f"en el archivo de entrada."
            ),
            "anomalias": analisis["anomalias"],
            "encoding": analisis["encoding"],
            "juridicas": juridicas,
            "contenido_juridicas": _lineas_origen(juridicas),
            "csv_juridicas": _csv_juridicas(juridicas),
        }

    filas.sort(key=lambda x: x["cbu"])

    registros = []
    total_imp = 0
    no_maestro = []
    seq = 1

    for item in filas:
        info = buscar_en_maestro(maestro, item["cbu"])
        if info and info["nombre"]:
            nombre = formatear_nombre_desde_maestro(info["nombre"])
        else:
            nombre = formatear_nombre_desde_input(item["nombre"])
            no_maestro.append((item["cbu"], item["nombre"]))

        registros.append(build_221(seq, item, fecha_proceso, cfg))
        registros.append(build_222(seq, nombre, cfg))
        registros.append(build_223(seq, cfg))
        registros.append(build_224(seq, cfg))
        total_imp += item["importe"]
        seq += 1

    total_lineas = 1 + len(registros) + 1
    header = build_header(fecha_proceso, cfg)
    footer = build_footer(total_imp, len(filas), total_lineas, cfg)

    lines = [header] + registros + [footer]
    contenido = "\r\n".join(lines) + "\r\n"

    return {
        "contenido":      contenido,
        "beneficiarios":  len(filas),
        "total_lineas":   total_lineas,
        "importe_total":  total_imp / 100,
        "fecha":          fecha_proceso,
        "no_maestro":     no_maestro,
        "nombre_archivo": "BBVA_generado.txt",
        "maestro_registros": len(maestro),
        # ── Diagnóstico del archivo de origen ──
        "encoding":       analisis["encoding"],
        "anomalias":      [a for a in analisis["anomalias"]
                           if a["cbu"].startswith(otros_bancos.COD_BBVA)],
        "corridas":       sum(1 for r in aptos if r["corrido"]),
        "descartadas":    len(analisis["descartados"]),
        # ── Personas jurídicas separadas ──
        "leidos_017":            len(regs),
        "juridicas":             juridicas,
        "importe_juridicas":     sum(r["importe"] for r in juridicas) / 100,
        "contenido_juridicas":   _lineas_origen(juridicas),
        "csv_juridicas":         _csv_juridicas(juridicas),
        "nombre_arch_juridicas": "BBVA_excluidas_juridicas.txt",
        "nombre_csv_juridicas":  "BBVA_excluidas_juridicas.csv",
    }
