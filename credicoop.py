"""
Conversor: Otros Bancos (input) → Formato Credicoop (output)

Formato de salida (13 campos separados por ';'):
  CBU(22) ; Monto(15) ; TitularDesc(100) ; TitularCui(11) ;
  TipoCuentaDestino(3) ; CuentaDestino(18) ; CuentaDestinoPBF(18) ;
  esCuentaPropia(1) ; TipoPersona(1) ; Concepto(3) ; DescConcepto(12) ;
  Observaciones(60) ; Email(99) + \\x1e

Cada línea termina con \\r\\n.
Genera archivos divididos en chunks de 200 registros + ZIP consolidado.
"""

import io
import re
import zipfile
from openpyxl import load_workbook


# ══════════════════════════════════════════════════════════════
# CONFIGURACIÓN POR DEFECTO
# ══════════════════════════════════════════════════════════════
DEFAULTS = {
    "tipo_cuenta":  "CAP",          # Caja de Ahorro Pesos
    "es_propia":    "N",            # No es cuenta propia
    "tipo_persona": "J",            # Jurídica
    "observaciones": "COLEGIO MEDICO",
    "chunk_size":   200,            # Registros por archivo
}

RECORD_SEP = "\x1e"
ENC = "latin1"


# ══════════════════════════════════════════════════════════════
# FUNCIONES DE APOYO
# ══════════════════════════════════════════════════════════════
def cargar_maestro(file_bytes: bytes) -> dict:
    """Retorna dict CBU -> {email, cuenta}."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Hoja1"]
    d = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cbu_raw = row[1]
        if cbu_raw is None:
            continue
        cbu = str(cbu_raw).strip()
        mail = row[6]
        mail = "" if (mail is None or mail == 0) else str(mail).strip()
        nro_cta = row[3]
        nro_cta_clean = re.sub(r"[^0-9]", "", str(nro_cta)) if nro_cta else ""
        d[cbu] = {"email": mail, "cuenta": nro_cta_clean}
    return d


def formato_monto(monto_10_str: str) -> str:
    """Convierte monto de 10 dígitos (2 decimales implícitos) a '000000032052,74'."""
    valor = int(monto_10_str.strip())
    entero = valor // 100
    decimal = valor % 100
    return f"{entero:012d},{decimal:02d}"


def construir_linea(cbu, monto_fmt, nombre_100, cuit_11, concepto_3,
                    desc_12, observ_60, cuenta_18, email_99, cfg):
    campos = [
        cbu.ljust(22)[:22],
        monto_fmt,
        nombre_100.ljust(100)[:100],
        cuit_11.ljust(11)[:11],
        cfg["tipo_cuenta"],
        cuenta_18.zfill(18)[:18],
        cuenta_18.zfill(18)[:18],
        cfg["es_propia"],
        cfg["tipo_persona"],
        concepto_3.ljust(3)[:3],
        desc_12.ljust(12)[:12],
        observ_60.ljust(60)[:60],
        email_99.ljust(99)[:99] + RECORD_SEP,
    ]
    return ";".join(campos)


def validar_estructura(lineas: list) -> list:
    esperado = [22, 15, 100, 11, 3, 18, 18, 1, 1, 3, 12, 60, 100]
    errores = []
    for i, ln in enumerate(lineas, 1):
        campos = ln.split(";")
        if len(campos) != 13:
            errores.append(f"Línea {i}: {len(campos)} campos (se esperan 13)")
            continue
        for j, (c, e) in enumerate(zip(campos, esperado)):
            if len(c) != e:
                errores.append(f"Línea {i} campo {j+1}: largo {len(c)} (esperado {e})")
    return errores


# ══════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════
def convertir(input_bytes: bytes, maestro_bytes: bytes, config: dict = None) -> dict:
    """
    Ejecuta la conversión completa.

    Parámetros:
        input_bytes:   bytes del archivo Otros_Bancos (input).txt
        maestro_bytes: bytes del archivo Maestro_de_datos_Credicoop.xlsx
        config:        dict con valores de configuración (opcional)

    Retorna dict con:
        "archivos":       list de dicts {nombre, contenido_bytes}
        "zip_bytes":      bytes del ZIP consolidado
        "convertidas":    int
        "descartadas":    int
        "sin_datos":      list de CBUs sin datos en maestro
        "errores":        list de errores de validación
        "nombre_zip":     str
    """
    cfg = {**DEFAULTS, **(config or {})}
    chunk_size = cfg["chunk_size"]

    maestro = cargar_maestro(maestro_bytes)
    texto = input_bytes.decode(ENC)

    salida = []
    descartadas = 0
    sin_datos = []

    for ln in texto.splitlines():
        if not ln.strip():
            continue
        if ln.startswith("285") or ln.startswith("017"):
            descartadas += 1
            continue

        cbu       = ln[0:22]
        monto_raw = ln[22:32]
        nombre    = ln[32:132]
        concepto  = ln[132:135]
        desc_conc = ln[135:147]
        cuit      = ln[198:209]

        info = maestro.get(cbu.strip())
        if info:
            email  = info["email"]
            cuenta = info["cuenta"]
        else:
            email  = ""
            cuenta = ""
            sin_datos.append(cbu.strip())

        monto_fmt = formato_monto(monto_raw)

        linea = construir_linea(
            cbu=cbu, monto_fmt=monto_fmt, nombre_100=nombre,
            cuit_11=cuit, concepto_3=concepto, desc_12=desc_conc,
            observ_60=cfg["observaciones"], cuenta_18=cuenta,
            email_99=email, cfg=cfg,
        )
        salida.append(linea)

    if not salida:
        return {
            "archivos": [],
            "zip_bytes": b"",
            "convertidas": 0,
            "descartadas": descartadas,
            "sin_datos": sin_datos,
            "errores": ["No se generaron registros de salida."],
            "nombre_zip": "Credicoop_consolidado.zip",
        }

    errores = validar_estructura(salida)

    # Dividir en chunks
    archivos = []
    for idx, start in enumerate(range(0, len(salida), chunk_size), 1):
        chunk = salida[start:start + chunk_size]
        nombre = f"Credicoop_consolidado_parte{idx}.txt"
        contenido = ("\r\n".join(chunk) + "\r\n").encode(ENC)
        archivos.append({"nombre": nombre, "contenido_bytes": contenido})

    # Generar ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as z:
        for a in archivos:
            z.writestr(a["nombre"], a["contenido_bytes"])
    zip_bytes = zip_buffer.getvalue()

    return {
        "archivos":     archivos,
        "zip_bytes":    zip_bytes,
        "convertidas":  len(salida),
        "descartadas":  descartadas,
        "sin_datos":    sin_datos,
        "errores":      errores,
        "nombre_zip":   "Credicoop_consolidado.zip",
        "maestro_registros": len(maestro),
    }
