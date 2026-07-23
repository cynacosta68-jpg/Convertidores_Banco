"""
Conversor: Otros Bancos (input) → Formato BBVA (output)

- Extrae del archivo de input los registros cuyo CBU pertenece al BBVA (017)
  usando el parser robusto de `otros_bancos` (soporta líneas corridas)
- Cruza con el Maestro Excel para obtener nombres formateados
- Genera archivo TXT de 250 caracteres/línea según modelo BBVA

Estructura del archivo de SALIDA (cada línea = 250 chars + CRLF):
  ┌─ 211 (Cabecera)       × 1
  ├─ 221 (Detalle dato)   ┐
  ├─ 222 (Detalle nombre) │ × N beneficiarios
  ├─ 223 (Detalle prov.)  │
  ├─ 224 (Detalle ref.)   ┘
  └─ 291 (Pie / totales)  × 1
"""

import io
from datetime import datetime
from openpyxl import load_workbook

try:
    from convertidores import otros_bancos
except ImportError:                      # ejecución con módulos planos
    import otros_bancos


# ══════════════════════════════════════════════════════════════
# CONFIGURACIÓN POR DEFECTO
# ══════════════════════════════════════════════════════════════
DEFAULTS = {
    "servicio":      "001261",
    "cod_banco":     "0017",
    "cta_debito":    "0097740103302327",
    "concepto":      "HONORARIOS",
    "moneda":        "ARS",
    "flag":          "0",
    "banco_nombre":  "FRANCES",
    "empresa":       "COLEGIO MEDICO DEL SUR DEL CHUBUT",
    "cod_extra":     "20",
    "provincia":     "CHUBUT",
    "referencia":    "LIQUIDACION",
    "tipo_cuenta":   "1",       # 1=CA, 0=CC
}

LINE_LEN = 250


# ══════════════════════════════════════════════════════════════
# FUNCIONES DE APOYO
# ══════════════════════════════════════════════════════════════
def pad(texto: str, n: int) -> str:
    return texto[:n].ljust(n)


def cargar_maestro(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    maestro = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cbu_raw = str(row[idx["CBU"]] or "").strip()
        if not cbu_raw:
            continue
        maestro[cbu_raw] = {
            "nombre":     str(row[idx["NombreApellido"]] or "").strip(),
            "nro_cuenta": str(row[idx["NroCuenta"]] or "").strip(),
            "tipo_cobro": str(row[idx["TipoCobro"]] or "").strip(),
        }
    wb.close()
    return maestro


def parsear_fila_input(linea: str) -> dict:
    """
    Adaptador: delega en el parser robusto de `otros_bancos`.

    Antes esta función cortaba la línea por posiciones fijas, lo que
    devolvía importes truncados y CUITs corridos cuando el registro no
    medía exactamente 209 caracteres. Ver otros_bancos.py.
    """
    reg = otros_bancos.parsear_linea(linea)
    return {
        "cbu":      reg["cbu"],
        "importe":  reg["importe"],
        "nombre":   reg["nombre"],
        "tipo_doc": reg["tipo_doc"],
        "cuil":     reg["cuit"],
        "avisos":   reg["avisos"],
    }


def formatear_nombre_desde_maestro(nombre_maestro: str) -> str:
    partes = nombre_maestro.split()
    if not partes:
        return nombre_maestro
    apellidos = []
    nombres = []
    for p in partes:
        if p.isupper() and len(p) > 1:
            apellidos.append(p)
        else:
            nombres.append(p)
    if apellidos and nombres:
        nombre_fmt = " ".join(n.capitalize() if n.islower() else n for n in nombres)
        return f"{' '.join(apellidos)}, {nombre_fmt}"
    return nombre_maestro


def formatear_nombre_desde_input(nombre_input: str) -> str:
    if "," in nombre_input:
        ape, nom = nombre_input.split(",", 1)
        nom_title = " ".join(w.capitalize() for w in nom.strip().split())
        return f"{ape.strip()}, {nom_title}"
    return nombre_input


def buscar_en_maestro(maestro: dict, cbu: str):
    cbu_limpio = cbu.strip()
    if cbu_limpio in maestro:
        return maestro[cbu_limpio]
    for k, v in maestro.items():
        if k.strip() == cbu_limpio:
            return v
    return None


# ══════════════════════════════════════════════════════════════
# GENERADORES DE REGISTROS
# ══════════════════════════════════════════════════════════════
def build_header(fecha: str, cfg: dict) -> str:
    h = (
        "211"
        + cfg["servicio"]
        + fecha + fecha
        + cfg["cod_banco"]
        + pad(cfg["cta_debito"], 16)
        + pad(cfg["concepto"], 10)
        + cfg["moneda"]
        + cfg["flag"]
        + pad(cfg["banco_nombre"], 12)
        + pad(cfg["empresa"], 36)
        + pad(cfg["cod_extra"], 3)
    )
    return pad(h, LINE_LEN)


def build_221(seq: int, item: dict, fecha: str, cfg: dict) -> str:
    return pad(
        "221" + cfg["servicio"] + "  "
        + str(seq).zfill(18) + "    "
        + cfg["tipo_cuenta"]
        + item["cbu"]
        + str(item["importe"]).zfill(25)
        + "      " + fecha
        + item["cuil"].zfill(15),
        LINE_LEN
    )


def build_222(seq: int, nombre: str, cfg: dict) -> str:
    return pad(
        "222" + cfg["servicio"] + "  "
        + str(seq).zfill(18) + "    "
        + nombre,
        LINE_LEN
    )


def build_223(seq: int, cfg: dict) -> str:
    return pad(
        "223" + cfg["servicio"] + "  "
        + str(seq).zfill(18) + "    "
        + " " * 36 + cfg["provincia"],
        LINE_LEN
    )


def build_224(seq: int, cfg: dict) -> str:
    return pad(
        "224" + cfg["servicio"] + "  "
        + str(seq).zfill(18) + "    "
        + cfg["referencia"],
        LINE_LEN
    )


def build_footer(total_importe: int, cant_benef: int, total_lineas: int, cfg: dict) -> str:
    return pad(
        "291" + cfg["servicio"]
        + str(total_importe).zfill(15)
        + str(cant_benef).zfill(8) + "00"
        + str(total_lineas).zfill(8),
        LINE_LEN
    )


# ══════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════
def convertir(input_bytes: bytes, maestro_bytes: bytes,
              fecha_proceso: str = None, config: dict = None) -> dict:
    """
    Ejecuta la conversión completa.

    Parámetros:
        input_bytes:   bytes del archivo Otros_Bancos (input).txt
        maestro_bytes: bytes del archivo Maestro_datos_BBVA.xlsx
        fecha_proceso: str AAAAMMDD (default: hoy)
        config:        dict con valores de configuración (opcional)

    Retorna dict con:
        "contenido":      str con el archivo generado
        "beneficiarios":  int
        "total_lineas":   int
        "importe_total":  float
        "fecha":          str
        "no_maestro":     list de tuplas (cbu, nombre)
        "nombre_archivo": str sugerido
    """
    cfg = {**DEFAULTS, **(config or {})}
    if not fecha_proceso:
        fecha_proceso = datetime.today().strftime("%Y%m%d")

    maestro = cargar_maestro(maestro_bytes)

    # ── Lectura robusta del archivo Otros Bancos ──────────────────
    # Se filtra por código de banco del CBU (017 = BBVA). El parser
    # detecta la codificación real y reacomoda las líneas corridas.
    regs, analisis = otros_bancos.registros_para(
        input_bytes, incluir=[otros_bancos.COD_BBVA]
    )

    filas = [{
        "cbu":      r["cbu"],
        "importe":  r["importe"],
        "nombre":   r["nombre"],
        "tipo_doc": r["tipo_doc"],
        "cuil":     r["cuit"],
        "avisos":   r["avisos"],
    } for r in regs]

    if not filas:
        return {
            "contenido": "",
            "beneficiarios": 0,
            "error": f"No se encontraron registros del banco {otros_bancos.COD_BBVA} "
                     f"en el archivo de entrada.",
            "anomalias": analisis["anomalias"],
            "encoding": analisis["encoding"],
        }

    filas.sort(key=lambda x: x["cbu"])

    registros = []
    total_imp = 0
    no_maestro = []
    seq = 1

    for item in filas:
        info = buscar_en_maestro(maestro, item["cbu"])
        if info and info["nombre"]:
            nombre = formatear_nombre_desde_maestro(info["nombre"])
        else:
            nombre = formatear_nombre_desde_input(item["nombre"])
            no_maestro.append((item["cbu"], item["nombre"]))

        registros.append(build_221(seq, item, fecha_proceso, cfg))
        registros.append(build_222(seq, nombre, cfg))
        registros.append(build_223(seq, cfg))
        registros.append(build_224(seq, cfg))
        total_imp += item["importe"]
        seq += 1

    total_lineas = 1 + len(registros) + 1
    header = build_header(fecha_proceso, cfg)
    footer = build_footer(total_imp, len(filas), total_lineas, cfg)

    lines = [header] + registros + [footer]
    contenido = "\r\n".join(lines) + "\r\n"

    return {
        "contenido":      contenido,
        "beneficiarios":  len(filas),
        "total_lineas":   total_lineas,
        "importe_total":  total_imp / 100,
        "fecha":          fecha_proceso,
        "no_maestro":     no_maestro,
        "nombre_archivo": "BBVA_generado.txt",
        "maestro_registros": len(maestro),
        # ── Diagnóstico del archivo de origen ──
        "encoding":       analisis["encoding"],
        "anomalias":      [a for a in analisis["anomalias"]
                           if a["cbu"].startswith(otros_bancos.COD_BBVA)],
        "corridas":       sum(1 for r in regs if r["corrido"]),
        "descartadas":    len(analisis["descartados"]),
    }
