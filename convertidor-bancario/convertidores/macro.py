"""
Conversor: PS.txt → Formato Banco MACRO (output)

Formato PS.txt (TAB):     0000000 | CUIT | (vacío) | CBU | 0 | importe | 0
Formato MACRO (TAB):      0000000 | CUIT | NombreApellido | NroCuenta | CBU | importe | fecha(DDMMAA)
"""

import io
from datetime import datetime
from openpyxl import load_workbook


# ══════════════════════════════════════════════════════════════
# FUNCIONES DE APOYO
# ══════════════════════════════════════════════════════════════
def cargar_maestro(file_bytes: bytes) -> dict:
    """Devuelve dict: CBU -> (NombreApellido, NroCuenta)."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx_cbu = headers.index("CBU")
    idx_nom = headers.index("NombreApellido")
    idx_cta = headers.index("NroCuenta")

    maestro = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[idx_cbu] is None:
            continue
        cbu = str(row[idx_cbu]).strip()
        nombre = str(row[idx_nom]).strip().upper() if row[idx_nom] else ""
        cuenta = str(row[idx_cta]).strip() if row[idx_cta] is not None else ""
        maestro[cbu] = (nombre, cuenta)
    return maestro


def formatear_importe(valor) -> str:
    return f"{float(str(valor).replace(',', '.')):.2f}"


# ══════════════════════════════════════════════════════════════
# FUNCIÓN PRINCIPAL
# ══════════════════════════════════════════════════════════════
def convertir(input_bytes: bytes, maestro_bytes: bytes,
              fecha_ddmmaa: str = None) -> dict:
    """
    Ejecuta la conversión de PS.txt a formato MACRO.

    Parámetros:
        input_bytes:   bytes del archivo PS.txt
        maestro_bytes: bytes del archivo Maestro_datos_MACRO.xlsx
        fecha_ddmmaa:  str DDMMAA (default: hoy)

    Retorna dict con:
        "contenido":       str con el archivo generado
        "registros":       int cantidad procesados
        "no_encontrados":  list de dicts {linea, cuit, cbu}
        "nombre_archivo":  str sugerido
        "fecha":           str
    """
    if not fecha_ddmmaa:
        fecha_ddmmaa = datetime.now().strftime("%d%m%y")

    maestro = cargar_maestro(maestro_bytes)

    # Intentar decodificar con utf-8, si falla con latin-1
    try:
        texto = input_bytes.decode("utf-8")
    except UnicodeDecodeError:
        texto = input_bytes.decode("latin-1")

    salida = []
    no_encontrados = []
    lineas_invalidas = 0

    for nlinea, linea in enumerate(texto.splitlines(), 1):
        if not linea.strip():
            continue
        campos = [c.strip() for c in linea.split("\t")]
        if len(campos) < 6:
            lineas_invalidas += 1
            continue

        cero    = campos[0]
        cuit    = campos[1]
        cbu     = campos[3].strip()
        importe = formatear_importe(campos[5])

        if cbu not in maestro:
            no_encontrados.append({"linea": nlinea, "cuit": cuit, "cbu": cbu})
            continue

        nombre, nro_cuenta = maestro[cbu]
        fila = "\t".join([cero, cuit, nombre, nro_cuenta, cbu, importe, fecha_ddmmaa])
        salida.append(fila)

    contenido = "\n".join(salida) + "\n" if salida else ""

    return {
        "contenido":       contenido,
        "registros":       len(salida),
        "no_encontrados":  no_encontrados,
        "lineas_invalidas": lineas_invalidas,
        "nombre_archivo":  "salida_MACRO.txt",
        "fecha":           fecha_ddmmaa,
        "maestro_registros": len(maestro),
    }
